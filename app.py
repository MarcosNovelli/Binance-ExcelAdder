from time import time
from typing import List
from binance.client import Client
import config
from openpyxl import Workbook, load_workbook
import time
import datetime

Trades_Filename = config.TRADES_FILENAME
client = Client(config.API_KEY, config.API_SECRET, tld='com')

def trade_retriever(symbol:str, id:int) -> List:
    trade = client.futures_account_trades(symbol=symbol, fromId=id)
    return trade[0]

def turn_unix_into_date(unix_time:int):
    unix_time = int(str(unix_time)[:-3])
    time = str(datetime.datetime.fromtimestamp(unix_time))
    return time

def return_last_registered_tradeId(symbol:str) -> str:
    """
    Needed: Symbol to look up last trade registered on excel
    Returns: ID of last registered trade on excel
    """
    try:
        wb = load_workbook(Trades_Filename)
        ws = wb[symbol]
        id:str = 0
        last_table_col:str = str(((ws.tables.items()[0][1])).split(':')[1][0])
        last_table_row:str = str(((ws.tables.items()[0][1])).split(':')[1][1:])
        
        while not id:
            id:str = ws[last_table_col + last_table_row].value
            last_table_row = str(int(last_table_row) - 1)
            if id == 'Last Order ID':
                print(f"{symbol}'s table is empty")
                id = 0
                break 

        wb.save(Trades_Filename)
    
    except Exception as e:
        print("an exception occured in return_last_registered_tradeId - {}".format(e))
    
    return id

def get_last_closing_trade(symbol) -> List:
    """
    Needed: Symbol
    Returns: List of all the information on las closing trade, even if trade is divided
        into multiple orders 
    """
    try:
        n:int = 2
        ct_type:str = "Closing Order"

        # Get info on last closing order
        closing_order = client.futures_account_trades(symbol=symbol, limit=1)

        symbol:str = closing_order[0]['symbol']
        side:str = closing_order[0]['side']
        ct_quantity:float = float(closing_order[0]['qty'])
        ct_price:float = float(closing_order[0]['price'])
        ct_pnl:float = float(closing_order[0]['realizedPnl'])
        ct_commission:float = float(closing_order[0]['commission'])
        ct_id = closing_order[0]['id']
        ct_time = closing_order[0]['time']
        
        while client.futures_account_trades(symbol=symbol, limit=n)[0]['side'] == side:
            new_trade = client.futures_account_trades(symbol=symbol, limit=n)
            ct_price = float(new_trade[0]['price'])
            ct_quantity += float(new_trade[0]['qty'])
            ct_pnl += float(new_trade[0]['realizedPnl'])
            ct_commission += float(new_trade[0]['commission'])
            ct_id = new_trade[0]['id']

            n += 1
        ct_trade = [ct_time, symbol, side, ct_type, ct_price, ct_quantity, ct_pnl, ct_commission, ct_id]
    
    except Exception as e:
        print("an exception occured in get_last_closing_trade - {}".format(e))

    return ct_trade


def get_opening_trade(symbol:str, ct_side:str, quantity:float) -> List:
    """
    Needed: Symbol, side to search for and quantity the orders should accumulate to 
    Returns: Opening order 
    """
    try:
        n:int = 2
        new_side = ct_side
        ot_type = "Opening Order"
        
        while new_side == ct_side :
            new_order = client.futures_account_trades(symbol=symbol, limit=n)
            new_side = new_order[0]['side']
            n += 1

        side:str = new_side
        ot_quantity:float = float(new_order[0]['qty'])
        ot_price:float = float(new_order[0]['price'])
        ot_pnl:float = float(new_order[0]['realizedPnl'])
        ot_commission:float = float(new_order[0]['commission'])
        ot_id = new_order[0]['id']
        ot_time = new_order[0]['time']
        
        while ot_quantity < quantity:
            new_order = client.futures_account_trades(symbol=symbol, limit=n)
            ot_pnl += float(new_order[0]['realizedPnl'])
            ot_commission += float(new_order[0]['commission'])
            ot_quantity += float(new_order[0]['qty'])
            n += 1
        
        ot_trade = [ot_time, symbol, side, ot_type, ot_price, ot_quantity, ot_pnl, ot_commission, ot_id]
    
    except Exception as e:
        print("an exception occured in get_opening_trade - {}".format(e))

    return ot_trade



def return_last_trade(symbol:str) -> List:
    '''
    Needed: A symbol to look up
    Returns: Last trade taken on symbol, both open and closing order, in two
        seperate lists
    '''
    
    #Get full closing trade
    ct_trade = get_last_closing_trade(symbol)
    
    ct_side  = ct_trade[1]
    quantity = ct_trade[4]

    #Get full opening order
    ot_trade = get_opening_trade(symbol, ct_side, quantity)
    
    return  ot_trade, ct_trade


def add_to_excel(data:dict):
    """
    Needed: Trade data, trade has to be closed, and symbol in second list index
    """
    wb = load_workbook(Trades_Filename)
    ws = wb[data[0][1]]
    
    # Devuelve el nombre del table y el rango que ocupa
    last_row = int((str(ws.tables.items()[0][1])).split(':')[1][1:])
  
    for col in range(len(data)):
        column = 1
        #considerando que siempre arranca en A1, encontrar manera de resolver esto (puede ser usando los del rango que ocupa el table)    
        for n in range(len(data[0])):
            _ = ws.cell(column=column, row=last_row + 1, value=data[col][n])
            column += 1
        last_row += 1
    
    wb.save(Trades_Filename)


def add_all_trades_to_excel(symbol:str, fromId=0):
    '''
    Needed: Symbol to search for trades and ID if
        wanting to start from a certain trade
    '''
    trade_list = []
    all_trades = client.futures_account_trades(symbol=symbol, fromId=fromId)
    i = 0
    
    while i < len(all_trades) :
        side:str = all_trades[i]['side']
        quantity:float = float(all_trades[i]['qty'])
        price:float = float(all_trades[i]['price'])
        pnl:float = float(all_trades[i]['realizedPnl'])
        commission:float = float(all_trades[i]['commission'])
        id:int = all_trades[i]['id']
        time = all_trades[i]['time']
        
        if pnl == 0:
            type = 'Opening Order'
        else:
            type = 'Closing Order'

        # Transformar time a formato YYYY-MM-DD HH:MM:SS
        time = turn_unix_into_date(time)
        
        #Saca info del next trade
        if i < len(all_trades) - 1:
            next_pnl = float(all_trades[i + 1]['realizedPnl'])
            
            if next_pnl == 0:
                next_type = 'Opening Order'
            else:
                next_type = 'Closing Order'
            
            #If next trade same type then add everything
            while next_type == type and i < len(all_trades) - 1:
                i += 1
                quantity += float(all_trades[i]['qty'])
                pnl += float(all_trades[i]['realizedPnl'])
                commission += float(all_trades[i]['commission'])
                id = all_trades[i]['id']
                
                if i < len(all_trades) - 1:
                    next_pnl = float(all_trades[i + 1]['realizedPnl'])
                    
                    if next_pnl == 0:
                        next_type = 'Opening Order'
                    else:
                        next_type = 'Closing Order'                

        i += 1

        

        data = [time, symbol, side, type, price, quantity, pnl, commission, id]
        trade_list.append(data)

    if not trade_list:
        print(f"{symbol}'s trades are all up to date")
        return        

    try:
        add_to_excel(trade_list, Trades_Filename)
    except Exception as e:
        print("An exception has occured in adding the trades to excel - {}".format(e))
    print(f"All {symbol}'s trades have been added")

def add_missing_trades(symbol:str):
    """
    Needed: Symbol to add al missing trades to excel
    """
    last_trade = return_last_registered_tradeId(symbol, Trades_Filename) + 1

    add_all_trades_to_excel(symbol, last_trade)
    return

add_missing_trades("ETHUSDT", Trades_Filename)
add_missing_trades("SOLUSDT", Trades_Filename)
add_missing_trades("MATICUSDT", Trades_Filename)
add_missing_trades("XRPUSDT", Trades_Filename)
add_missing_trades("BTCUSDT", Trades_Filename)
add_missing_trades("BNBUSDT", Trades_Filename)


